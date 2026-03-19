"""
Export ScoringResults to an Excel spreadsheet for human review,
and import the reviewed spreadsheet back into ReviewRecords.

Columns:
  student_id | student_name | total_score | total_max | pct | confidence
  | needs_review | breakdown_json | uncertain_parts_json | llm_reasoning
  | reviewer_override_score | reviewer_notes | approved
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from models import ReviewRecord, ScoringResult

# Yellow fill for rows that need human review
_REVIEW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
_HEADER_FONT = Font(bold=True)

COLUMNS = [
    "student_id",
    "student_name",
    "bb_user_id",
    "assignment_id",
    "total_score",
    "total_max",
    "pct",
    "confidence",
    "needs_review",
    "breakdown_json",
    "uncertain_parts_json",
    "llm_reasoning",
    # --- human fills these ---
    "reviewer_override_score",
    "reviewer_notes",
    "approved",
]


def export(results: list[ScoringResult], path: Path) -> None:
    """Write scoring results to an Excel file."""
    import json

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"

    # Header row
    for col, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = _HEADER_FONT

    # Freeze header
    ws.freeze_panes = "A2"

    for row_idx, r in enumerate(results, start=2):
        row_data = [
            r.student_id,
            r.student_name,
            r.bb_user_id,
            r.assignment_id,
            r.total_score,
            r.total_max,
            r.pct,
            r.confidence,
            "YES" if r.needs_review else "NO",
            json.dumps([b.model_dump() for b in r.breakdown], ensure_ascii=False),
            json.dumps([u.model_dump() for u in r.uncertain_parts], ensure_ascii=False),
            r.llm_reasoning,
            "",   # reviewer_override_score
            "",   # reviewer_notes
            "NO", # approved — reviewer changes to YES
        ]
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col, value=value)

        # Highlight rows that need review
        if r.needs_review:
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col).fill = _REVIEW_FILL

    # Auto-width for readability
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(path)


def load_reviewed(path: Path) -> list[ReviewRecord]:
    """Read a reviewed spreadsheet back into ReviewRecord objects."""
    import json

    from models import CriterionScore, ScoringResult, UncertainPart

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    idx = {name: i for i, name in enumerate(headers)}

    records: list[ReviewRecord] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["student_id"]]:
            continue

        breakdown = [CriterionScore(**b) for b in json.loads(row[idx["breakdown_json"]] or "[]")]
        uncertain = [UncertainPart(**u) for u in json.loads(row[idx["uncertain_parts_json"]] or "[]")]

        result = ScoringResult(
            student_id=str(row[idx["student_id"]]),
            student_name=str(row[idx["student_name"]]),
            bb_user_id=str(row[idx["bb_user_id"]] or "") if "bb_user_id" in idx else "",
            assignment_id=str(row[idx["assignment_id"]]),
            total_score=float(row[idx["total_score"]]),
            total_max=float(row[idx["total_max"]]),
            confidence=float(row[idx["confidence"]]),
            needs_review=row[idx["needs_review"]] == "YES",
            breakdown=breakdown,
            uncertain_parts=uncertain,
            llm_reasoning=str(row[idx["llm_reasoning"]] or ""),
        )

        override_raw = row[idx["reviewer_override_score"]]
        override = float(override_raw) if override_raw not in (None, "") else None
        approved = str(row[idx["approved"]]).strip().upper() == "YES"

        records.append(ReviewRecord(
            result=result,
            reviewer_override_score=override,
            reviewer_notes=str(row[idx["reviewer_notes"]] or ""),
            approved=approved,
        ))

    return records
