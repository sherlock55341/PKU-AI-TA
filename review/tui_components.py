"""
Components for the interactive TUI reviewer.
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
import termios
import tty
from pathlib import Path
from typing import Any

import openpyxl
from rich.console import Console
from rich.console import Group
from rich.live import Live
from rich.panel import Panel
from rich.prompt import Prompt
from rich.table import Table
from rich.text import Text

# Enable readline for better line editing (arrow keys, history, etc.)
try:
    import readline
    # Optional: Configure readline for better behavior
    readline.parse_and_bind('set editing-mode emacs')
except ImportError:
    # Try pyreadline3 for Windows (pip install pyreadline3)
    try:
        import pyreadline as readline  # type: ignore
    except ImportError:
        # No readline available, but input() still works on Windows
        # Windows console has basic line editing support built-in
        pass


def prompt_text(prompt_label: str, default: str = "", console: Console | None = None, allow_interrupt: bool = False) -> str:
    """Prompt for text input using Python's built-in input() for better IME and cursor support.

    This is better than rich.prompt.Prompt for:
    - Chinese/Japanese/Korean input (IME support)
    - Left/right arrow keys for cursor movement
    - Proper backspace handling with multi-byte characters

    If allow_interrupt is True, KeyboardInterrupt will be raised instead of returning default.
    """
    # Print the prompt on its own line first to avoid readline cursor position issues
    if console:
        console.print(f"{prompt_label}", end="")
        if default:
            console.print(f" [dim](default: {default})[/dim]", end="")
        console.print()
    else:
        print(f"{prompt_label}{f' (default: {default})' if default else ''}")

    # Then just use a simple "> " prompt for input()
    try:
        result = input("> ")
        return result.strip() if result.strip() != "" else default
    except EOFError:
        print()
        return default
    except KeyboardInterrupt:
        print()
        if allow_interrupt:
            raise
        return default


def prompt_choice(prompt_label: str, choices: list[str], default: str | None = None, console: Console | None = None) -> str:
    """Prompt for a choice from a list using rich.prompt.Prompt.

    This is fine for simple menu selections where no text editing is needed.
    """
    return Prompt.ask(prompt_label, choices=choices, default=default)


def find_submission_file(submissions_dir: Path, student_id: str, student_name: str) -> Path | None:
    """Find submission file for a student by matching student_id in filename."""
    if not submissions_dir.exists():
        return None
    for assignment_dir in submissions_dir.iterdir():
        if not assignment_dir.is_dir() or assignment_dir.name.startswith(".`"):
            continue
        for f in assignment_dir.iterdir():
            if f.is_file() and not f.name.startswith(".") and student_id in f.name:
                return f
    return None


def open_file(filepath: Path, console: Console | None = None) -> None:
    """Open file with default system viewer. Cross-platform support."""
    import shutil
    filepath_str = str(filepath)
    try:
        if sys.platform == "darwin":  # macOS
            subprocess.run(["open", filepath_str], check=False)
        elif sys.platform == "win32":  # Windows
            os.startfile(filepath_str)  # type: ignore
        else:  # Linux / Unix variants
            # Try openers in order of preference, checking if they exist first
            for opener in ["xdg-open", "gio", "gnome-open", "kde-open"]:
                if shutil.which(opener):
                    try:
                        subprocess.run([opener, filepath_str], check=False,
                                      stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                        break
                    except Exception:
                        continue
    except Exception as e:
        if console:
            console.print(f"[yellow]Warning: Could not open file ({e})[/yellow]")
            console.print(f"[dim]Please open manually: {filepath_str}[/dim]")

def needs_review_check(row_data: dict) -> bool:
    """Check if a student needs review even if already approved."""
    # Check if score is perfect (consider override if present)
    total_max = float(row_data.get("total_max", 100) or 100)
    override_score = row_data.get("reviewer_override_score")
    if override_score is not None and override_score != "":
        try:
            current_score = float(override_score)
        except (ValueError, TypeError):
            current_score = float(row_data.get("total_score", 0) or 0)
    else:
        current_score = float(row_data.get("total_score", 0) or 0)
    has_notes = bool(str(row_data.get("reviewer_notes") or "").strip())
    # Needs review if not perfect score and no notes, even if approved
    return current_score < total_max and not has_notes

def load_review_data(
    scores: Path,
    needs_review_only: bool,
    all_students: bool,
) -> tuple[Any, dict, list[tuple[int, dict]]]:
    """Load student data from scores spreadsheet and filter based on review status."""
    wb = openpyxl.load_workbook(scores)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    idx = {name: i for i, name in enumerate(headers)}

    rows: list[tuple[int, dict]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[idx["student_id"]]:
            continue
        row_data = {name: row[i] if i < len(row) else None for name, i in idx.items()}
        if needs_review_only and row_data.get("needs_review") != "YES":
            continue
        # Skip approved students only if they don't need review for missing notes
        is_approved = str(row_data.get("approved", "")).upper() == "YES"
        if not all_students and is_approved and not needs_review_check(row_data):
            continue
        rows.append((row_idx, row_data))
    
    return wb, idx, rows

def auto_approve_students(ws: Any, idx: dict, console: Console, *, approve_all_safe: bool = False) -> bool:
    """Auto-approve students who do not need review.

    By default, only perfect-score submissions are auto-approved.
    When approve_all_safe is True, all submissions with needs_review=NO are auto-approved.
    """
    auto_approved = 0
    modified = False
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[idx["student_id"]]:
            continue
        row_data = {name: row[i] if i < len(row) else None for name, i in idx.items()}
        total_score = float(row_data.get("total_score", 0) or 0)
        total_max = float(row_data.get("total_max", 100) or 100)
        needs_review = row_data.get("needs_review") == "YES"
        already_approved = str(row_data.get("approved", "")).upper() == "YES"

        should_approve = (
            not needs_review
            and not already_approved
            and (approve_all_safe or total_score >= total_max)
        )
        if should_approve:
            student_name = row_data.get("student_name", "")
            student_id = row_data.get("student_id", "")
            console.print(f"  [dim]Auto-approving:[/dim] {student_name} ({student_id}) — {total_score}/{total_max}")
            ws.cell(row=row_idx, column=idx["approved"] + 1, value="YES")
            auto_approved += 1
            modified = True

    if auto_approved > 0:
        if approve_all_safe:
            console.print(f"[green]Auto-approved {auto_approved} student(s) with needs_review=NO.[/green]")
        else:
            console.print(f"[green]Auto-approved {auto_approved} student(s) with perfect scores.[/green]")
    
    return modified

class ReviewSession:
    """Manages the state of an interactive review session."""
    def __init__(self, scores: Path, needs_review_only: bool, all_students: bool):
        self.scores = scores
        self.wb, self.idx, self.rows = load_review_data(scores, needs_review_only, all_students)
        self.ws = self.wb.active
        self.current_idx = 0
        self.modified_rows: dict[int, dict] = {}
        self.modified = False

    def get_current_row(self) -> tuple[int, dict] | None:
        if 0 <= self.current_idx < len(self.rows):
            row_idx, row_data_original = self.rows[self.current_idx]
            row_data = self.modified_rows.get(row_idx, row_data_original.copy())
            return row_idx, row_data
        return None

    def next_student(self):
        self.current_idx += 1

    def prev_student(self):
        if self.current_idx > 0:
            self.current_idx -= 1

    def update_row(self, row_idx: int, row_data: dict):
        self.modified_rows[row_idx] = row_data
        self.modified = True

    def save_changes(self):
        if self.modified:
            self.wb.save(self.scores)

def handle_approve(session: ReviewSession, row_idx: int, row_data: dict, console: Console) -> bool:
    """Handle the 'approve' action. Returns True if approved, False if cancelled."""
    total_max = float(row_data.get("total_max", 100) or 100)
    override_score = row_data.get("reviewer_override_score")
    if override_score is not None and override_score != "":
        try:
            current_score = float(override_score)
        except (ValueError, TypeError):
            current_score = float(row_data.get("total_score", 0) or 0)
    else:
        current_score = float(row_data.get("total_score", 0) or 0)
    is_perfect = current_score >= total_max
    has_notes = bool(str(row_data.get("reviewer_notes") or "").strip())

    if not is_perfect and not has_notes:
        console.print("\n[yellow]Score is not 100%. Please add reviewer notes.[/yellow]")
        current_notes = str(row_data.get("reviewer_notes") or "")
        try:
            new_notes = prompt_text("[bold cyan]Enter reviewer notes[/bold cyan]", default=current_notes, console=console, allow_interrupt=True)
        except KeyboardInterrupt:
            console.print("\n[yellow]Approve cancelled.[/yellow]")
            return False
        if new_notes.strip():
            row_data["reviewer_notes"] = new_notes
            session.ws.cell(row=row_idx, column=session.idx["reviewer_notes"] + 1, value=new_notes)
            has_notes = True
            console.print("[green]Added reviewer notes.[/green]")
        else:
            console.print("[yellow]No notes added. You can still add notes later.[/yellow]")

    session.ws.cell(row=row_idx, column=session.idx["approved"] + 1, value="YES")
    row_data["approved"] = "YES"
    session.update_row(row_idx, row_data)
    console.print("[green]Marked as approved.[/green]")
    return True

def handle_notes(session: ReviewSession, row_idx: int, row_data: dict, console: Console) -> None:
    """Handle the 'notes' action."""
    current_notes = str(row_data.get("reviewer_notes") or "")
    console.print(f"\n[bold]Current notes:[/bold] {current_notes if current_notes else '(none)'}")
    try:
        new_notes = prompt_text("[bold cyan]Enter reviewer notes[/bold cyan]", default=current_notes, console=console, allow_interrupt=True)
    except KeyboardInterrupt:
        console.print("\n[yellow]Cancelled - notes not changed.[/yellow]")
        return
    row_data["reviewer_notes"] = new_notes
    session.ws.cell(row=row_idx, column=session.idx["reviewer_notes"] + 1, value=new_notes)
    session.update_row(row_idx, row_data)
    console.print("[green]Updated reviewer notes.[/green]")

def handle_edit(session: ReviewSession, row_idx: int, row_data: dict, breakdown: list, console: Console) -> list:
    """Handle the 'edit' action."""
    if not breakdown:
        console.print("[yellow]No breakdown data to edit[/yellow]")
        return breakdown
    breakdown, new_total, new_max = edit_breakdown(console, breakdown)
    # Update the row data
    row_data["breakdown_json"] = json.dumps(breakdown)
    row_data["total_score"] = new_total
    row_data["total_max"] = new_max
    row_data["pct"] = round((new_total / new_max) * 100, 1) if new_max > 0 else 0
    # Update Excel
    session.ws.cell(row=row_idx, column=session.idx["breakdown_json"] + 1, value=row_data["breakdown_json"])
    session.ws.cell(row=row_idx, column=session.idx["total_score"] + 1, value=new_total)
    session.ws.cell(row=row_idx, column=session.idx["total_max"] + 1, value=new_max)
    session.ws.cell(row=row_idx, column=session.idx["pct"] + 1, value=row_data["pct"])
    session.update_row(row_idx, row_data)
    console.print(f"[green]Updated total score: {new_total} / {new_max}[/green]")
    return breakdown


def handle_batch_edit(session: ReviewSession, row_idx: int, row_data: dict, breakdown: list, console: Console) -> list:
    """Handle batch editing of multiple criterion scores."""
    if not breakdown:
        console.print("[yellow]No breakdown data to edit[/yellow]")
        return breakdown
    new_breakdown = batch_edit_breakdown(console, breakdown)
    if new_breakdown is breakdown:
        return breakdown

    new_total, new_max = calculate_breakdown_totals(new_breakdown)
    row_data["breakdown_json"] = json.dumps(new_breakdown)
    row_data["total_score"] = new_total
    row_data["total_max"] = new_max
    row_data["pct"] = round((new_total / new_max) * 100, 1) if new_max > 0 else 0
    session.ws.cell(row=row_idx, column=session.idx["breakdown_json"] + 1, value=row_data["breakdown_json"])
    session.ws.cell(row=row_idx, column=session.idx["total_score"] + 1, value=new_total)
    session.ws.cell(row=row_idx, column=session.idx["total_max"] + 1, value=new_max)
    session.ws.cell(row=row_idx, column=session.idx["pct"] + 1, value=row_data["pct"])
    session.update_row(row_idx, row_data)
    console.print(f"[green]Updated total score: {new_total} / {new_max}[/green]")
    return new_breakdown


def handle_override(session: ReviewSession, row_idx: int, row_data: dict, console: Console) -> None:
    """Handle the 'override' action. Sets override score and stays on current student."""
    current_override = row_data.get("reviewer_override_score", "")
    if current_override:
        console.print(f"\n[bold]Current override:[/bold] {current_override}")
    new_score = Prompt.ask("[bold cyan]Enter override score[/bold cyan]", default=str(current_override) if current_override else "")
    if new_score:
        try:
            score_val = float(new_score)
            row_data["reviewer_override_score"] = score_val
            session.ws.cell(row=row_idx, column=session.idx["reviewer_override_score"] + 1, value=score_val)
            session.update_row(row_idx, row_data)
            console.print(f"[green]Override score set to {score_val}. Use (a) to approve.[/green]")
        except ValueError:
            console.print("[red]Invalid score[/red]")

def calculate_breakdown_totals(breakdown: list) -> tuple[float, float]:
    """Return total awarded and max points for a breakdown list."""
    total = sum(float(b.get("points_awarded", 0)) for b in breakdown)
    max_total = sum(float(b.get("points_max", 0)) for b in breakdown)
    return total, max_total


def apply_batch_scores(breakdown: list, raw: str) -> tuple[list, list[str]]:
    """Apply batch score edits.

    Supported formats:
    - "1=8,2=6,5=0" updates explicit criterion numbers.
    - "8,6,,10" updates criteria in order; blanks keep existing scores.
    """
    updated = [dict(item) for item in breakdown]
    errors: list[str] = []
    text = raw.strip()
    if not text:
        return updated, errors

    parts = [part.strip() for part in text.replace("\n", ",").split(",")]
    explicit = any("=" in part for part in parts if part)
    if explicit:
        for part in parts:
            if not part:
                continue
            if "=" not in part:
                errors.append(f"Expected N=score, got: {part}")
                continue
            left, right = [p.strip() for p in part.split("=", 1)]
            try:
                index = int(left) - 1
            except ValueError:
                errors.append(f"Invalid criterion number: {left}")
                continue
            _apply_score_at_index(updated, index, right, errors)
        return updated, errors

    for index, score_text in enumerate(parts):
        if not score_text:
            continue
        _apply_score_at_index(updated, index, score_text, errors)
    return updated, errors


def batch_edit_breakdown(console: Console, breakdown: list) -> list:
    """Edit multiple scores using a keyboard-driven table when available."""
    if sys.stdin.isatty():
        return navigate_edit_breakdown(console, breakdown)
    return prompt_batch_scores(console, breakdown)


def prompt_batch_scores(console: Console, breakdown: list) -> list:
    """Prompt once for multiple score edits and return an updated breakdown."""
    console.print()
    console.print(Panel(
        "[bold]Batch Edit Criterion Scores[/bold]\n"
        "Formats: [cyan]1=8,2=6,5=0[/cyan] or [cyan]8,6,,10[/cyan]\n"
        "Blank entries keep current scores.",
        border_style="magenta",
    ))
    _print_breakdown_score_table(console, breakdown)
    current_total, current_max = calculate_breakdown_totals(breakdown)
    console.print(f"\n[bold]Current Total:[/bold] {current_total} / {current_max}")
    raw = prompt_text("[bold magenta]Batch scores[/bold magenta]", default="", console=console)
    if not raw.strip():
        console.print("[dim]No batch edits applied.[/dim]")
        return breakdown

    updated, errors = apply_batch_scores(breakdown, raw)
    if errors:
        for error in errors:
            console.print(f"[red]{error}[/red]")
        console.print("[yellow]Batch edit cancelled; no scores changed.[/yellow]")
        return breakdown
    return updated


def navigate_edit_breakdown(console: Console, breakdown: list) -> list:
    """Use arrow keys to choose a criterion and edit scores in-place."""
    updated = [dict(item) for item in breakdown]
    selected = 0
    message = ""

    with Live(_navigate_editor_renderable(updated, selected, message), console=console, refresh_per_second=12, transient=True) as live:
        while True:
            live.update(_navigate_editor_renderable(updated, selected, message))
            key = _read_key()
            message = ""
            if key in ("q", "Q"):
                return updated
            if key in ("\x1b", "\x03"):
                console.print("[yellow]Batch edit cancelled; no scores changed.[/yellow]")
                return breakdown
            if key in ("up", "k"):
                selected = max(0, selected - 1)
                continue
            if key in ("down", "j"):
                selected = min(len(updated) - 1, selected + 1)
                continue
            if key in ("\r", "\n", "e"):
                live.stop()
                item = updated[selected]
                max_val = float(item.get("points_max", 0))
                current = str(item.get("points_awarded", 0))
                console.print(f"\n[bold]Editing #{selected + 1}:[/bold] {item.get('criterion', '')}")
                new_score = Prompt.ask(f"New score (0-{max_val})", default=current)
                errors: list[str] = []
                _apply_score_at_index(updated, selected, new_score, errors)
                message = f"[red]{errors[0]}[/red]" if errors else f"[green]Updated #{selected + 1} to {updated[selected]['points_awarded']}[/green]"
                live.start(refresh=True)


def _navigate_editor_renderable(breakdown: list, selected: int, message: str = "") -> Group:
    header = Panel(
        "[bold]Batch Edit Criterion Scores[/bold]\n"
        "Use [cyan]↑/↓[/cyan] or [cyan]k/j[/cyan] to move, [cyan]Enter[/cyan] to edit, "
        "[cyan]q[/cyan] to save and return, [cyan]Esc[/cyan] to cancel.",
        border_style="magenta",
    )
    table = Table()
    table.add_column("", width=2)
    table.add_column("#", style="dim", justify="right")
    table.add_column("Criterion", style="cyan")
    table.add_column("Awarded", justify="right")
    table.add_column("Max", justify="right")
    for i, item in enumerate(breakdown):
        awarded = float(item.get("points_awarded", 0))
        max_p = float(item.get("points_max", 0))
        pointer = ">" if i == selected else ""
        style = "bold reverse" if i == selected else ("red" if awarded < max_p else "green")
        table.add_row(pointer, str(i + 1), str(item.get("criterion", "")), Text(f"{awarded}", style=style), f"{max_p}")
    total, max_total = calculate_breakdown_totals(breakdown)
    footer = Text.from_markup(f"[bold]Current Total:[/bold] {total} / {max_total}")
    if message:
        return Group(header, table, footer, Text.from_markup(message))
    return Group(header, table, footer)


def _read_key() -> str:
    fd = sys.stdin.fileno()
    old_settings = termios.tcgetattr(fd)
    try:
        tty.setraw(fd)
        ch = sys.stdin.read(1)
        if ch == "\x1b":
            rest = sys.stdin.read(2)
            if rest == "[A":
                return "up"
            if rest == "[B":
                return "down"
            return "\x1b"
        return ch
    finally:
        termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)


def _apply_score_at_index(breakdown: list, index: int, score_text: str, errors: list[str]) -> None:
    if index < 0 or index >= len(breakdown):
        errors.append(f"Criterion number out of range: {index + 1}")
        return
    try:
        score_val = float(score_text)
    except ValueError:
        errors.append(f"Invalid score for criterion {index + 1}: {score_text}")
        return
    max_val = float(breakdown[index].get("points_max", 0))
    if not 0 <= score_val <= max_val:
        errors.append(f"Score for criterion {index + 1} must be between 0 and {max_val}")
        return
    breakdown[index]["points_awarded"] = score_val


def _print_breakdown_score_table(console: Console, breakdown: list) -> None:
    table = Table()
    table.add_column("#", style="dim", justify="right")
    table.add_column("Criterion", style="cyan")
    table.add_column("Awarded", justify="right", style="green")
    table.add_column("Max", justify="right")
    for i, item in enumerate(breakdown, start=1):
        awarded = float(item.get("points_awarded", 0))
        max_p = float(item.get("points_max", 0))
        style = "red" if awarded < max_p else "green"
        table.add_row(str(i), str(item.get("criterion", "")), Text(f"{awarded}", style=style), f"{max_p}")
    console.print(table)


def edit_breakdown(console: Console, breakdown: list) -> tuple[list, float, float]:
    """Interactive editor for breakdown. Returns (new_breakdown, new_total, new_max)."""
    while True:
        console.print()
        console.print(Panel("[bold]Edit Criterion Scores[/bold]\n"
                           "Enter the number of the criterion to edit, or 'd' when done",
                           border_style="magenta"))

        _print_breakdown_score_table(console, breakdown)

        # Calculate and show current total
        current_total, current_max = calculate_breakdown_totals(breakdown)
        console.print(f"\n[bold]Current Total:[/bold] {current_total} / {current_max}")
        console.print()

        choice = Prompt.ask("[bold magenta]Criterion # to edit, or [d]one[/bold magenta]",
                           choices=[str(i) for i in range(1, len(breakdown) + 1)] + ["d", "done"],
                           default="d")

        if choice.lower() in ("d", "done"):
            new_total, new_max = calculate_breakdown_totals(breakdown)
            return breakdown, new_total, new_max

        # Edit selected criterion
        idx = int(choice) - 1
        item = breakdown[idx]
        console.print(f"\n[bold]Editing:[/bold] {item.get('criterion', '')}")
        console.print(f"  Current: {item.get('points_awarded', 0)} / {item.get('points_max', 0)}")
        console.print(f"  Reasoning: {item.get('reasoning', '')}")

        new_score = Prompt.ask(f"  New score (0-{item.get('points_max', 0)})",
                              default=str(item.get("points_awarded", 0)))
        try:
            score_val = float(new_score)
            max_val = float(item.get("points_max", 0))
            if 0 <= score_val <= max_val:
                item["points_awarded"] = score_val
                console.print(f"[green]Updated to {score_val} / {max_val}[/green]")
            else:
                console.print(f"[red]Score must be between 0 and {max_val}[/red]")
        except ValueError:
            console.print("[red]Invalid number[/red]")
