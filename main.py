"""
PKU AI Teaching Assistant CLI

Commands:
  ta grade   --course <id> --column <id> --rubric <file> [--whitelist a,b,c] [--out scores.xlsx] [--verbose] [--resume] [--lang en|zh]
  ta review  [--scores scores.xlsx] [--submissions submissions/] [--needs-review] [--all]
  ta submit  --course <id> --column <id> --scores <reviewed.xlsx> [--dry-run]
"""
from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Annotated, Optional
from time import time

import typer
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, BarColumn, TaskProgressColumn, TimeElapsedColumn, TextColumn

app = typer.Typer(help="PKU AI Teaching Assistant")
console = Console()


@app.command()
def grade(
    course: Annotated[str, typer.Option(help="Blackboard course ID, e.g. _12345_1")] = "",
    column: Annotated[str, typer.Option(help="Gradebook column (assignment) ID")] = "",
    rubric: Annotated[Path, typer.Option(help="Path to rubric file (any format the LLM supports)")] = Path("rubric.md"),
    whitelist: Annotated[str, typer.Option(help="Comma-separated student IDs to include; empty = all")] = "",
    out: Annotated[Path, typer.Option(help="Output Excel path")] = Path("scores.xlsx"),
    save_dir: Annotated[Optional[Path], typer.Option(help="Save submission files here for human review; default: submissions/")] = Path("submissions"),
    verbose: Annotated[bool, typer.Option("--verbose", "-v", help="Show intermediate scores for each student")] = False,
    resume: Annotated[bool, typer.Option("--resume", "-r", help="Resume from previous partial run (if any)")] = False,
    lang: Annotated[str, typer.Option(help="LLM prompt language: en or zh")] = "en",
) -> None:
    """Crawl submissions, score with LLM, export review spreadsheet.

    Press Ctrl-C to interrupt; partial results will be saved to the output file
    and can be resumed with --resume.
    """
    from threading import Lock

    from auth.iaaa import get_session
    from config import settings
    from crawler.pku_homework import PKUHomeworkCrawler
    from review.spreadsheet import export
    from scorer.llm import score_submission
    from models import ScoringResult

    # Checkpoint save/load using Excel format
    checkpoint_path = out
    all_results: list[ScoringResult] = []
    processed_ids: set[str] = set()
    save_lock = Lock()

    def save_checkpoint() -> None:
        """Save current progress to output Excel file."""
        with save_lock:
            if all_results:
                export(all_results, checkpoint_path)

    def load_checkpoint() -> list[ScoringResult]:
        """Load previous progress from output Excel file (if exists and --resume is set)."""
        if resume and checkpoint_path.exists():
            try:
                from review.spreadsheet import load_reviewed
                records = load_reviewed(checkpoint_path)
                results = [r.result for r in records]
                console.print(f"[bold cyan]Resuming from checkpoint:[/bold cyan] {len(results)} previously processed result(s)")
                return results
            except Exception as e:
                console.print(f"[yellow]Warning: Could not load checkpoint: {e}[/yellow]")
        return []

    # Load checkpoint if resuming
    if resume:
        all_results = load_checkpoint()
        processed_ids = {r.student_id for r in all_results}

    # Resolve config — CLI args override .env
    course_id = course or settings.course_id
    if not course_id:
        console.print("[red]Error:[/red] --course is required (or set COURSE_ID in .env)")
        raise typer.Exit(1)

    whitelist_ids: set[str] = (
        {s.strip() for s in whitelist.split(",") if s.strip()}
        if whitelist
        else settings.whitelist_ids
    )

    if not rubric.exists():
        console.print(f"[red]Error:[/red] Rubric file not found: {rubric}")
        raise typer.Exit(1)

    rubric_text = rubric.read_text(encoding="utf-8")

    console.print("[bold]Step 1/3:[/bold] Authenticating with PKU IAAA…")
    client = get_session()

    crawler = PKUHomeworkCrawler(client, course_id, whitelist_ids)

    if column:
        # column here is expected to be gradeBookPK (numeric), e.g. "423829"
        columns = [{"gradeBookPK": column, "name": column, "id": f"_{column}_1"}]
    else:
        console.print("[bold]Step 1b:[/bold] Fetching assignment list…")
        columns = crawler.fetch_assignments()
        console.print(f"  Found {len(columns)} assignment(s).")

    interrupted = False
    start_time = time()

    try:
        for col in columns:
            grade_book_pk = col.get("gradeBookPK") or col["id"].strip("_").split("_")[0]
            col_title = col.get("name") or col["id"]
            console.print(f"\n[bold]Step 2/3:[/bold] Fetching submissions for [cyan]{col_title}[/cyan]…")

            submissions = crawler.fetch_submissions(grade_book_pk, col_title)
            if not submissions:
                console.print("  No submissions found.")
                continue

            # Filter out already processed submissions if resuming
            if processed_ids:
                submissions = [s for s in submissions if s.student_id not in processed_ids]
                if not submissions:
                    console.print("  All submissions already processed.")
                    continue
                console.print(f"  {len(submissions)} submission(s) remaining to process")

            if save_dir:
                _save_submissions(submissions, save_dir, col_title)
                console.print(f"  Saved files → [cyan]{save_dir / col_title}[/cyan]")

            total_submissions = len(submissions)
            console.print(f"  Scoring {total_submissions} submission(s) with LLM (threads={settings.ta_threads}, lang={lang})…")
            console.print(f"  [dim]Press Ctrl-C to interrupt — progress will be saved[/dim]")

            # Use transient=False for verbose mode so results stay on screen
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                TaskProgressColumn(),
                TimeElapsedColumn(),
                TextColumn("[dim]ETA: {task.fields[eta]}"),
                console=console, transient=not verbose,
            ) as progress:
                task = progress.add_task("  Scoring", total=total_submissions, eta="calculating...")
                completed_count = 0

                with ThreadPoolExecutor(max_workers=settings.ta_threads) as executor:
                    futures = {executor.submit(score_submission, sub, rubric_text, lang): sub for sub in submissions}
                    for future in as_completed(futures):
                        sub = futures[future]
                        try:
                            result = future.result()
                            all_results.append(result)
                            processed_ids.add(result.student_id)
                            completed_count += 1

                            # Calculate ETA
                            if completed_count >= 2:
                                elapsed = time() - start_time
                                avg_time_per = elapsed / completed_count
                                remaining = (total_submissions - completed_count) * avg_time_per
                                if remaining < 60:
                                    eta_str = f"{remaining:.0f}s"
                                elif remaining < 3600:
                                    eta_str = f"{remaining/60:.1f}m"
                                else:
                                    eta_str = f"{remaining/3600:.1f}h"
                                progress.update(task, eta=eta_str)
                            else:
                                progress.update(task, eta="...")

                            # Save checkpoint after each result for safety
                            save_checkpoint()

                            if verbose:
                                # Show verbose output for each student
                                needs_review = result.needs_review
                                color = "yellow" if needs_review else "green"
                                status = "NEEDS_REVIEW" if needs_review else "OK"
                                console.print(
                                    f"  [{color}]{result.student_id:12s}[/] {result.student_name:10s} "
                                    f"→ {result.total_score:3.0f}/{result.total_max:3.0f} ({result.pct:3.0f}%) "
                                    f"[{color}]{status}[/]"
                                )
                        except Exception as e:
                            console.print(f"  [red]Error scoring {sub.student_id}:[/red] {e}")
                        finally:
                            progress.advance(task)

    except KeyboardInterrupt:
        console.print("\n[yellow]Interrupted by user[/yellow]")
        interrupted = True
        if all_results:
            console.print(f"[yellow]Saving {len(all_results)} partial result(s)...[/yellow]")
            save_checkpoint()
            console.print(f"[cyan]Checkpoint saved to {checkpoint_path}[/cyan]")
            console.print(f"[cyan]Resume later with: --resume[/cyan]")
        raise typer.Exit(1)

    if not all_results:
        console.print("[yellow]No results to export.[/yellow]")
        raise typer.Exit(0)

    console.print(f"\n[bold]Step 3/3:[/bold] Exporting {len(all_results)} result(s) → [cyan]{out}[/cyan]")
    export(all_results, out)

    needs_review = sum(1 for r in all_results if r.needs_review)
    console.print(
        f"\n[green]Done.[/green] {needs_review}/{len(all_results)} submission(s) flagged for review "
        f"(highlighted in yellow in the spreadsheet)."
    )
    console.print(f"Review the spreadsheet, set [bold]approved[/bold] = YES, then run [bold]ta submit[/bold].")


@app.command()
def submit(
    course: Annotated[str, typer.Option(help="Blackboard course ID")] = "",
    column: Annotated[str, typer.Option(help="Gradebook column (assignment) ID")] = "",
    scores: Annotated[Path, typer.Option(help="Reviewed Excel spreadsheet")] = Path("scores.xlsx"),
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Print what would be submitted without posting")] = False,
) -> None:
    """Submit approved scores from the reviewed spreadsheet back to course.pku.edu.cn."""
    from auth.iaaa import get_session
    from config import settings
    from review.spreadsheet import load_reviewed
    from submitter.blackboard import submit_scores

    course_id = course or settings.course_id
    if not course_id or not column:
        console.print("[red]Error:[/red] Both --course and --column are required.")
        raise typer.Exit(1)
    # BB REST API needs _423829_1 format; accept bare numeric gradeBookPK too
    col_id = column if column.startswith("_") else f"_{column}_1"

    if not scores.exists():
        console.print(f"[red]Error:[/red] Scores file not found: {scores}")
        raise typer.Exit(1)

    records = load_reviewed(scores)
    approved_count = sum(1 for r in records if r.approved)
    console.print(f"Loaded {len(records)} record(s), {approved_count} approved.")

    if approved_count == 0:
        console.print("[yellow]Nothing to submit — no records marked approved.[/yellow]")
        raise typer.Exit(0)

    console.print("[bold]Authenticating with PKU IAAA…[/bold]")
    client = get_session()

    submit_scores(client, course_id, col_id, records, dry_run=dry_run)


@app.command()
def review(
    scores: Annotated[Path, typer.Option(help="Excel spreadsheet to review")] = Path("scores.xlsx"),
    submissions: Annotated[Path, typer.Option(help="Directory with submission files")] = Path("submissions"),
    rubric: Annotated[Path, typer.Option(help="Path to rubric file to open during review")] = Path("rubric.md"),
    needs_review_only: Annotated[bool, typer.Option("--needs-review", "-n", help="Only review students marked needs_review=YES")] = False,
    all_students: Annotated[bool, typer.Option("--all", "-a", help="Review all students (including already approved)")] = False,
) -> None:
    """Interactive TUI for reviewing submissions one by one.

    Shows score breakdown, opens submission file, and lets you approve or override scores.
    Press 'e' to edit individual criterion scores, 'r' to open the rubric.
    """
    import json
    import os
    import subprocess
    import sys

    import openpyxl
    from rich.panel import Panel
    from rich.prompt import Confirm, Prompt, IntPrompt
    from rich.table import Table
    from rich.text import Text

    def find_submission_file(submissions_dir: Path, student_id: str, student_name: str) -> Path | None:
        """Find submission file for a student by matching student_id in filename."""
        if not submissions_dir.exists():
            return None
        for assignment_dir in submissions_dir.iterdir():
            if not assignment_dir.is_dir() or assignment_dir.name.startswith("."):
                continue
            for f in assignment_dir.iterdir():
                if f.is_file() and not f.name.startswith(".") and student_id in f.name:
                    return f
        return None

    def open_file(filepath: Path) -> None:
        """Open file with default system viewer. Cross-platform support."""
        filepath_str = str(filepath)
        try:
            if sys.platform == "darwin":  # macOS
                subprocess.run(["open", filepath_str], check=False)
            elif sys.platform == "win32":  # Windows
                os.startfile(filepath_str)  # type: ignore
            else:  # Linux / Unix variants
                for opener in ["xdg-open", "gio", "gnome-open", "kde-open"]:
                    try:
                        subprocess.run([opener, filepath_str], check=False,
                                      stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                        break
                    except (subprocess.SubprocessError, FileNotFoundError):
                        continue
        except Exception as e:
            console.print(f"[yellow]Warning: Could not open file ({e})[/yellow]")
            console.print(f"[dim]Please open manually: {filepath_str}[/dim]")

    def display_student(row_data: dict, row_idx: int, total: int, breakdown: list | None = None) -> list:
        """Display student info and score breakdown using Rich. Returns the breakdown list."""
        console.print()
        console.print(Panel.fit(
            f"[bold cyan]Student {row_idx}/{total}[/bold cyan]",
            title="Review Progress",
            border_style="blue"
        ))
        info_table = Table(show_header=False, box=None)
        info_table.add_row("[bold]Student ID:[/]", str(row_data["student_id"]))
        info_table.add_row("[bold]Name:[/]", str(row_data["student_name"]))
        info_table.add_row("[bold]Score:[/]", f"{row_data['total_score']} / {row_data['total_max']} ({row_data['pct']}%)")
        info_table.add_row("[bold]Confidence:[/]", f"{row_data['confidence']}")
        info_table.add_row("[bold]Needs review:[/]", "[yellow]YES[/yellow]" if row_data["needs_review"] == "YES" else "NO")
        info_table.add_row("[bold]Current approved:[/]", "[green]YES[/green]" if str(row_data.get("approved", "")).upper() == "YES" else "[red]NO[/red]")
        console.print(Panel(info_table, title="Student Info", border_style="cyan"))

        if breakdown is None:
            try:
                breakdown = json.loads(row_data.get("breakdown_json", "[]"))
            except json.JSONDecodeError:
                breakdown = []
                console.print("[yellow]Warning: Could not parse breakdown_json[/yellow]")

        if breakdown:
            bd_table = Table(title="Score Breakdown")
            bd_table.add_column("#", style="dim", justify="right")
            bd_table.add_column("Criterion", style="cyan")
            bd_table.add_column("Awarded", justify="right", style="green")
            bd_table.add_column("Max", justify="right")
            bd_table.add_column("Reasoning", style="dim")
            for i, item in enumerate(breakdown, start=1):
                awarded = float(item.get("points_awarded", 0))
                max_p = float(item.get("points_max", 0))
                style = "red" if awarded < max_p else "green"
                bd_table.add_row(
                    str(i),
                    str(item.get("criterion", "")),
                    Text(f"{awarded}", style=style),
                    f"{max_p}",
                    str(item.get("reasoning", ""))
                )
            console.print(Panel(bd_table, border_style="green"))

        try:
            uncertain = json.loads(row_data.get("uncertain_parts_json", "[]"))
            if uncertain:
                uc_table = Table(title="Uncertain Parts")
                uc_table.add_column("Description", style="yellow")
                uc_table.add_column("Suggested Score", justify="right")
                for item in uncertain:
                    uc_table.add_row(
                        str(item.get("description", "")),
                        f"{item.get('suggested_score', 0)} / {item.get('suggested_max', 0)}"
                    )
                console.print(Panel(uc_table, border_style="yellow"))
        except json.JSONDecodeError:
            pass

        if row_data.get("llm_reasoning"):
            console.print(Panel(
                Text(str(row_data["llm_reasoning"]), style="dim"),
                title="LLM Reasoning",
                border_style="dim"
            ))

        return breakdown

    def edit_breakdown(breakdown: list) -> tuple[list, float, float]:
        """Interactive editor for breakdown. Returns (new_breakdown, new_total, new_max)."""
        while True:
            console.print()
            console.print(Panel("[bold]Edit Criterion Scores[/bold]\n"
                               "Enter the number of the criterion to edit, or 'd' when done",
                               border_style="magenta"))

            # Show current breakdown with numbers
            bd_table = Table()
            bd_table.add_column("#", style="dim", justify="right")
            bd_table.add_column("Criterion", style="cyan")
            bd_table.add_column("Awarded", justify="right", style="green")
            bd_table.add_column("Max", justify="right")
            for i, item in enumerate(breakdown, start=1):
                awarded = float(item.get("points_awarded", 0))
                max_p = float(item.get("points_max", 0))
                style = "red" if awarded < max_p else "green"
                bd_table.add_row(
                    str(i),
                    str(item.get("criterion", "")),
                    Text(f"{awarded}", style=style),
                    f"{max_p}"
                )
            console.print(bd_table)

            # Calculate and show current total
            current_total = sum(float(b.get("points_awarded", 0)) for b in breakdown)
            current_max = sum(float(b.get("points_max", 0)) for b in breakdown)
            console.print(f"\n[bold]Current Total:[/bold] {current_total} / {current_max}")
            console.print()

            choice = Prompt.ask("[bold magenta]Criterion # to edit, or [d]one[/bold magenta]",
                               choices=[str(i) for i in range(1, len(breakdown) + 1)] + ["d", "done"],
                               default="d")

            if choice.lower() in ("d", "done"):
                new_total = sum(float(b.get("points_awarded", 0)) for b in breakdown)
                new_max = sum(float(b.get("points_max", 0)) for b in breakdown)
                return breakdown, new_total, new_max

            # Edit selected criterion
            idx = int(choice) - 1
            item = breakdown[idx]
            console.print(f"\n[bold]Editing:[/bold] {item.get('criterion', '')}")
            console.print(f"  Current: {item.get('points_awarded', 0)} / {item.get('points_max', 0)}")
            console.print(f"  Reasoning: {item.get('reasoning', '')}")

            new_score = Prompt.ask(f"  New score (0-{item.get('points_max', 0)})",
                                  default=str(item.get("points_awarded", 0)))
            try:
                score_val = float(new_score)
                max_val = float(item.get("points_max", 0))
                if 0 <= score_val <= max_val:
                    item["points_awarded"] = score_val
                    console.print(f"[green]Updated to {score_val} / {max_val}[/green]")
                else:
                    console.print(f"[red]Score must be between 0 and {max_val}[/red]")
            except ValueError:
                console.print("[red]Invalid number[/red]")

    if not scores.exists():
        console.print(f"[red]Error:[/red] Scores file not found: {scores}")
        raise typer.Exit(1)

    # Show rubric info
    if rubric.exists():
        console.print(f"[bold blue]Rubric:[/bold blue] {rubric}")
    else:
        console.print(f"[yellow]Warning: Rubric file not found: {rubric}[/yellow]")

    console.print(f"[bold]Loading spreadsheet:[/bold] {scores}")
    wb = openpyxl.load_workbook(scores)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    idx = {name: i for i, name in enumerate(headers)}

    rows: list[tuple[int, dict]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[idx["student_id"]]:
            continue
        row_data = {name: row[i] if i < len(row) else None for name, i in idx.items()}
        if needs_review_only and row_data.get("needs_review") != "YES":
            continue
        if not all_students and str(row_data.get("approved", "")).upper() == "YES":
            continue
        rows.append((row_idx, row_data))

    if not rows:
        console.print("[yellow]No students to review with the current filters.[/yellow]")
        raise typer.Exit(0)

    console.print(f"[green]Found {len(rows)} student(s) to review.[/green]")
    modified = False

    for i, (row_idx, row_data) in enumerate(rows, start=1):
        # Load breakdown first
        try:
            breakdown = json.loads(row_data.get("breakdown_json", "[]"))
        except json.JSONDecodeError:
            breakdown = []

        while True:
            display_student(row_data, i, len(rows), breakdown)
            student_id = str(row_data["student_id"])
            student_name = str(row_data["student_name"])
            sub_file = find_submission_file(submissions, student_id, student_name)

            if sub_file:
                console.print(f"[bold blue]Submission:[/bold blue] {sub_file}")
            else:
                console.print("[yellow]Warning: No submission file found[/yellow]")
            if rubric.exists():
                console.print(f"[bold blue]Rubric:[/bold blue] {rubric}")
            console.print()

            action = Prompt.ask(
                "[bold cyan]Action[/bold cyan]",
                choices=["a", "approve", "s", "skip", "e", "edit", "o", "open", "r", "rubric", "ov", "override", "q", "quit"],
                default="skip"
            )

            if action in ("q", "quit"):
                console.print("[yellow]Quitting...[/yellow]")
                break

            if action in ("a", "approve"):
                ws.cell(row=row_idx, column=idx["approved"] + 1, value="YES")
                row_data["approved"] = "YES"
                modified = True
                console.print("[green]Marked as approved.[/green]")
                break

            elif action in ("e", "edit"):
                if not breakdown:
                    console.print("[yellow]No breakdown data to edit[/yellow]")
                    continue
                breakdown, new_total, new_max = edit_breakdown(breakdown)
                # Update the row data
                row_data["breakdown_json"] = json.dumps(breakdown)
                row_data["total_score"] = new_total
                row_data["total_max"] = new_max
                row_data["pct"] = round((new_total / new_max) * 100, 1) if new_max > 0 else 0
                # Update Excel
                ws.cell(row=row_idx, column=idx["breakdown_json"] + 1, value=row_data["breakdown_json"])
                ws.cell(row=row_idx, column=idx["total_score"] + 1, value=new_total)
                ws.cell(row=row_idx, column=idx["total_max"] + 1, value=new_max)
                ws.cell(row=row_idx, column=idx["pct"] + 1, value=row_data["pct"])
                modified = True
                console.print(f"[green]Updated total score: {new_total} / {new_max}[/green]")
                continue

            elif action in ("o", "open") and sub_file:
                open_file(sub_file)
                continue

            elif action in ("r", "rubric"):
                if rubric.exists():
                    open_file(rubric)
                else:
                    console.print(f"[yellow]Rubric file not found: {rubric}[/yellow]")
                continue

            elif action in ("ov", "override"):
                new_score = Prompt.ask("[bold cyan]Enter override score[/bold cyan]")
                if new_score:
                    try:
                        score_val = float(new_score)
                        ws.cell(row=row_idx, column=idx["reviewer_override_score"] + 1, value=score_val)
                        ws.cell(row=row_idx, column=idx["approved"] + 1, value="YES")
                        row_data["approved"] = "YES"
                        modified = True
                        console.print(f"[green]Set override score: {score_val} and marked as approved.[/green]")
                        break
                    except ValueError:
                        console.print("[red]Invalid score[/red]")
                continue

            else:
                console.print("[dim]Skipped.[/dim]")
                break

        if action in ("q", "quit"):
            break

    if modified:
        console.print()
        if Confirm.ask("[bold cyan]Save changes to spreadsheet?[/bold cyan]", default=True):
            wb.save(scores)
            console.print(f"[green]Saved to {scores}[/green]")
        else:
            console.print("[yellow]Changes not saved.[/yellow]")
    else:
        console.print("[dim]No changes made.[/dim]")


def _save_submissions(submissions: list, save_dir: Path, assignment_title: str) -> None:
    """Save each student's attachment file to save_dir/assignment_title/ for human review."""
    import re
    safe_title = re.sub(r'[^\w\u4e00-\u9fff\-]', '_', assignment_title)
    dest = save_dir / safe_title
    dest.mkdir(parents=True, exist_ok=True)
    for sub in submissions:
        for att in sub.attachments:
            ext = Path(att.filename).suffix or ""
            # Filename: studentId_studentName.ext  (e.g. 2300012345_张三.pdf)
            safe_name = re.sub(r'[^\w\u4e00-\u9fff]', '_', sub.student_name)
            filename = f"{sub.student_id}_{safe_name}{ext}"
            (dest / filename).write_bytes(att.data)


if __name__ == "__main__":
    app()
