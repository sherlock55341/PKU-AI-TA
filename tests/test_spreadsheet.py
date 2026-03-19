import tempfile
from pathlib import Path

import pytest

from models import CriterionScore, ReviewRecord, ScoringResult, UncertainPart
from review.spreadsheet import export, load_reviewed


def make_result(student_id="2100012345", needs_review=False, confidence=0.9) -> ScoringResult:
    return ScoringResult(
        student_id=student_id,
        student_name="Test Student",
        assignment_id="_1_1",
        total_score=85.0,
        total_max=100.0,
        confidence=confidence,
        needs_review=needs_review,
        breakdown=[
            CriterionScore(criterion="Correctness", points_awarded=40.0, points_max=50.0, reasoning="Good."),
        ],
        uncertain_parts=[
            UncertainPart(description="Unclear", suggested_score=5.0, suggested_max=10.0),
        ] if needs_review else [],
        llm_reasoning="Overall good.",
    )


class TestSpreadsheetRoundtrip:
    def test_export_creates_file(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "scores.xlsx"
            export([make_result()], path)
            assert path.exists()

    def test_roundtrip_basic(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "scores.xlsx"
            results = [make_result("2100012345"), make_result("2100012346")]
            export(results, path)

            # Simulate reviewer approving via openpyxl
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            approved_col = headers.index("approved") + 1
            for row in range(2, ws.max_row + 1):
                ws.cell(row, approved_col).value = "YES"
            wb.save(path)

            records = load_reviewed(path)
            assert len(records) == 2
            assert all(r.approved for r in records)
            assert records[0].result.student_id == "2100012345"

    def test_reviewer_override_preserved(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "scores.xlsx"
            export([make_result()], path)

            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            override_col = headers.index("reviewer_override_score") + 1
            approved_col = headers.index("approved") + 1
            notes_col = headers.index("reviewer_notes") + 1
            ws.cell(2, override_col).value = 70.0
            ws.cell(2, approved_col).value = "YES"
            ws.cell(2, notes_col).value = "Deducted for late submission"
            wb.save(path)

            records = load_reviewed(path)
            assert records[0].reviewer_override_score == 70.0
            assert records[0].final_score == 70.0
            assert records[0].reviewer_notes == "Deducted for late submission"

    def test_unapproved_rows_loaded_correctly(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "scores.xlsx"
            export([make_result()], path)
            records = load_reviewed(path)
            # Default approved value in exported sheet is "NO"
            assert records[0].approved is False

    def test_needs_review_row_present(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "scores.xlsx"
            export([make_result(needs_review=True)], path)

            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            nr_col = headers.index("needs_review") + 1
            assert ws.cell(2, nr_col).value == "YES"

    def test_breakdown_and_uncertain_survive_roundtrip(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "scores.xlsx"
            export([make_result(needs_review=True)], path)

            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            ws.cell(2, headers.index("approved") + 1).value = "YES"
            wb.save(path)

            records = load_reviewed(path)
            assert len(records[0].result.breakdown) == 1
            assert records[0].result.breakdown[0].criterion == "Correctness"
            assert len(records[0].result.uncertain_parts) == 1
            assert records[0].result.uncertain_parts[0].description == "Unclear"
